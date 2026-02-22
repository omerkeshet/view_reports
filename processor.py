from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
import os
import re
import zipfile
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# -----------------------------
# Helpers: month string
# -----------------------------
def previous_month_str(today: Optional[datetime] = None) -> str:
    """
    Returns previous month in MM/YYYY, matching your script behavior.
    """
    now = today or datetime.now()
    year = now.year
    month = now.month - 1
    if month == 0:
        month = 12
        year -= 1
    return f"{month:02d}/{year}"


# -----------------------------
# Stage 1: CLEAN (from your script)
# -----------------------------
def _find_table_start(df: pd.DataFrame) -> Optional[int]:
    # Same logic: find first row i such that row i has data and next 5 rows also have data
    for i in range(len(df) - 5):
        current_has_data = df.iloc[i].notna().any()
        next_5_have_data = all(df.iloc[i + 1 : i + 6].notna().any(axis=1))
        if current_has_data and next_5_have_data:
            return i
    return None


def _filter_rows(df: pd.DataFrame) -> pd.DataFrame:
    # Same logic: keep row 0 if >=4 filled. Other rows keep if current and previous both >=4.
    rows_to_keep = []
    for i in range(len(df)):
        current_filled = df.iloc[i].notna().sum()
        if i == 0:
            if current_filled >= 4:
                rows_to_keep.append(i)
        else:
            prev_filled = df.iloc[i - 1].notna().sum()
            if current_filled >= 4 and prev_filled >= 4:
                rows_to_keep.append(i)
    return df.iloc[rows_to_keep].reset_index(drop=True)


def _finalize_cleaned(df: pd.DataFrame) -> pd.DataFrame:
    # First row is header, rest is data
    if len(df) > 0:
        df.columns = df.iloc[0]
        df = df.iloc[1:].reset_index(drop=True)
        df.columns.name = None
        df = df.dropna(axis=1, how="all")  # drop empty columns
    return df


def clean_input_files_to_excels(
    uploaded_files: List[Tuple[str, bytes]],
) -> Dict[str, bytes]:
    """
    Input: list of (original_filename, file_bytes)
    Output: dict of cleaned Excel files (filename -> bytes)
    Replicates your per-sheet output naming:
      cleaned_{base}_{sheet}.xlsx
    Only keeps sheets with > 15 rows after cleaning.
    """
    outputs: Dict[str, bytes] = {}

    for original_name, file_bytes in uploaded_files:
        lower = original_name.lower()

        # Excel files
        if lower.endswith(".xlsx") or lower.endswith(".xls"):
            all_sheets = pd.read_excel(BytesIO(file_bytes), sheet_name=None, header=None)

            base_name, ext = os.path.splitext(original_name)
            ext = ext if ext else ".xlsx"

            for sheet_name, sheet_df in all_sheets.items():
                df = sheet_df.copy()

                table_start = _find_table_start(df)
                if table_start is None:
                    continue

                df = df.iloc[table_start:].reset_index(drop=True)
                df = _filter_rows(df)
                df = _finalize_cleaned(df)

                if len(df) <= 15:
                    continue

                safe_sheet = str(sheet_name).replace("/", "_").replace("\\", "_")
                out_name = f"cleaned_{base_name}_{safe_sheet}{ext}"

                out_buf = BytesIO()
                df.to_excel(out_buf, index=False)
                out_buf.seek(0)
                outputs[out_name] = out_buf.read()

        # CSV files (your script had a csv branch)
        elif lower.endswith(".csv"):
            df = pd.read_csv(BytesIO(file_bytes))
            df = df.dropna(how="all").dropna(axis=1, how="all").reset_index(drop=True)

            out_name = f"cleaned_{original_name}"
            out_buf = BytesIO()
            # Save CSV-as-CSV (keeping behavior consistent)
            df.to_csv(out_buf, index=False)
            out_buf.seek(0)
            outputs[out_name] = out_buf.read()

        else:
            # ignore unknown types
            continue

    return outputs


# -----------------------------
# Stage 2: MAP HOUSE NUMBER (from your script)
# -----------------------------
def normalize_dash(text: str) -> str:
    if not isinstance(text, str):
        return ""
    s = text.replace("–", "-").replace("—", "-")
    s = re.sub(r"\s*-\s*", " - ", s)  # normalize around dash
    s = re.sub(r"\s+", " ", s).strip()
    return s


def swap_dash(name: str) -> Optional[str]:
    """
    If name looks like "A - B", returns "B - A", else None.
    """
    if not isinstance(name, str):
        return None
    name_norm = normalize_dash(name)
    if " - " not in name_norm:
        return None
    left, right = name_norm.split(" - ", 1)
    return f"{right.strip()} - {left.strip()}"


def _decide_program_col(file_name: str) -> Optional[str]:
    file_name_lower = file_name.lower()
    file_name_upper = file_name.upper()

    if "פרטנר" in file_name:
        return "שם תוכן"
    if "יס" in file_name:
        return "תאור אירוע"
    if "screenil" in file_name_lower:
        return "Title Translation"
    if ("סטינג" in file_name) and ("vod" in file_name_lower or "VOD" in file_name):
        return "שם מלא"
    if "סטינג" in file_name:
        return "תוכן"
    return None


def _decide_house_key_col(file_name: str) -> Optional[str]:
    file_name_upper = file_name.upper()
    if "סלקום" in file_name:
        return "קוד מזהה"
    if ("הוט" in file_name) and ("ספריה" in file_name or "ספרייה" in file_name):
        return "מזהה ייחודי קשת"
    if "NEXT" in file_name_upper:
        return "מזהה ייחודי קשת NP"
    return None


def map_house_numbers(
    cleaned_excels: Dict[str, bytes],
    db_excel_bytes: bytes,
) -> Dict[str, bytes]:
    """
    Input: cleaned Excel files dict + DB file bytes
    Output: mapped_{cleaned_filename}.xlsx bytes dict
    """
    mapped_outputs: Dict[str, bytes] = {}

    db_df = pd.read_excel(BytesIO(db_excel_bytes), sheet_name=0)
    db_df.columns = db_df.columns.astype(str).str.strip()

    # Normalized DB lookups like your script
    db_df["norm_name"] = db_df["שם קשת טי וי"].apply(normalize_dash)

    # Longest 'שם תכנית בפלטפורמה' per normalized program name
    db_name_groups = (
        db_df.dropna(subset=["norm_name"])
        .assign(name_len=db_df["שם תכנית בפלטפורמה"].astype(str).str.len())
        .sort_values("name_len", ascending=False)
        .drop_duplicates(subset=["norm_name"])
    )
    name_to_house = db_name_groups.set_index("norm_name")["HOUSE_NUMBER"]
    name_to_platform = db_name_groups.set_index("norm_name")["שם תכנית בפלטפורמה"]

    # For house-number joins, prefer longest platform name per house number
    db_house_groups = (
        db_df.assign(name_len=db_df["שם תכנית בפלטפורמה"].astype(str).str.len())
        .sort_values("name_len", ascending=False)
        .drop_duplicates(subset=["HOUSE_NUMBER"])
    )
    house_to_platform = db_house_groups.set_index("HOUSE_NUMBER")["שם תכנית בפלטפורמה"]

    for cleaned_name, cleaned_bytes in cleaned_excels.items():
        # Only map cleaned EXCEL files; if a cleaned CSV exists, you can extend later if needed
        if not cleaned_name.lower().endswith((".xlsx", ".xls")):
            continue

        df = pd.read_excel(BytesIO(cleaned_bytes))
        df.columns = df.columns.astype(str).str.strip()

        program_col = _decide_program_col(cleaned_name)
        house_key_col = _decide_house_key_col(cleaned_name)

        output_name = f"mapped_{cleaned_name}"

        # Case 1: program-based mapping
        if program_col is not None:
            if program_col not in df.columns:
                # Save as-is
                out_buf = BytesIO()
                df.to_excel(out_buf, index=False)
                out_buf.seek(0)
                mapped_outputs[output_name] = out_buf.read()
                continue

            base_cols = ["HOUSE_NUMBER", "שם קשת טי וי", "שם תכנית בפלטפורמה"]

            db_by_keshet = (
                db_df[base_cols]
                .dropna(subset=["שם קשת טי וי"])
                .drop_duplicates(subset=["שם קשת טי וי"], keep="first")
            )
            db_by_platform = (
                db_df[base_cols]
                .dropna(subset=["שם תכנית בפלטפורמה"])
                .drop_duplicates(subset=["שם תכנית בפלטפורמה"], keep="first")
            )

            m1 = df.merge(
                db_by_keshet,
                left_on=program_col,
                right_on="שם קשת טי וי",
                how="left",
            )
            m2 = df.merge(
                db_by_platform,
                left_on=program_col,
                right_on="שם תכנית בפלטפורמה",
                how="left",
            )

            merged = df.copy()
            merged["HOUSE_NUMBER"] = m1["HOUSE_NUMBER"].combine_first(m2["HOUSE_NUMBER"])
            merged["שם תכנית בפלטפורמה"] = m1["שם תכנית בפלטפורמה"].combine_first(
                m2["שם תכנית בפלטפורמה"]
            )

            # Third method: swapped A - B -> B - A with normalized dashes
            no_match = merged["HOUSE_NUMBER"].isna()
            if no_match.any():
                orig_names = merged.loc[no_match, program_col]
                swapped_keys = orig_names.apply(swap_dash)
                norm_swapped = swapped_keys.apply(lambda x: normalize_dash(x) if isinstance(x, str) else "")

                swapped_house = norm_swapped.map(name_to_house)
                swapped_platform = norm_swapped.map(name_to_platform)

                fill_mask = no_match & swapped_house.notna()
                merged.loc[fill_mask, "HOUSE_NUMBER"] = swapped_house[fill_mask]
                merged.loc[fill_mask, "שם תכנית בפלטפורמה"] = swapped_platform[fill_mask]

            # Remove DB name columns + dedupe (as in your script)
            merged = merged.drop(columns=["שם תכנית בפלטפורמה", "שם קשת טי וי"], errors="ignore")
            merged = merged.drop_duplicates()

            out_buf = BytesIO()
            merged.to_excel(out_buf, index=False)
            out_buf.seek(0)
            mapped_outputs[output_name] = out_buf.read()
            continue

        # Case 2: house-key mapping (bring platform name)
        if house_key_col is not None:
            if house_key_col not in df.columns:
                out_buf = BytesIO()
                df.to_excel(out_buf, index=False)
                out_buf.seek(0)
                mapped_outputs[output_name] = out_buf.read()
                continue

            merged = df.merge(
                db_df[["HOUSE_NUMBER", "שם תכנית בפלטפורמה"]],
                left_on=house_key_col,
                right_on="HOUSE_NUMBER",
                how="left",
            )

            merged = merged.drop(columns=["HOUSE_NUMBER"], errors="ignore")
            merged = merged.drop(columns=["שם תכנית בפלטפורמה", "שם קשת טי וי"], errors="ignore")
            merged = merged.drop_duplicates()

            out_buf = BytesIO()
            merged.to_excel(out_buf, index=False)
            out_buf.seek(0)
            mapped_outputs[output_name] = out_buf.read()
            continue

        # Case 3: no rule matched -> copy as-is
        out_buf = BytesIO()
        df.to_excel(out_buf, index=False)
        out_buf.seek(0)
        mapped_outputs[output_name] = out_buf.read()

    return mapped_outputs


# -----------------------------
# Stage 3: FILL TEMPLATE (from your script)
# -----------------------------
def _decide_template_program_col(file_name: str) -> Optional[str]:
    file_name_lower = file_name.lower()
    file_name_upper = file_name.upper()

    if "פרטנר" in file_name:
        return "שם תוכן"
    if "יס" in file_name:
        return "תאור אירוע"
    if "screenil" in file_name_lower:
        return "Title Translation"
    if ("סטינג" in file_name) and ("vod" in file_name_lower or "VOD" in file_name):
        return "שם מלא"
    if "סטינג" in file_name:
        return "תוכן"
    if "סלקום" in file_name:
        return "שם פריט"
    if ("הוט" in file_name) and ("ספריה" in file_name or "ספרייה" in file_name):
        return "שם כותר"
    if "NEXT" in file_name_upper:
        return "שם כותר NP"
    return None


def _decide_mapped_house_col(file_name: str) -> str:
    file_name_upper = file_name.upper()
    if "סלקום" in file_name:
        return "קוד מזהה"
    if ("הוט" in file_name) and ("ספריה" in file_name or "ספרייה" in file_name):
        return "מזהה ייחודי קשת"
    if "NEXT" in file_name_upper:
        return "מזהה ייחודי קשת NP"
    return "HOUSE_NUMBER"


def _decide_viewers_col(file_name: str) -> Optional[str]:
    file_name_lower = file_name.lower()
    file_name_upper = file_name.upper()

    if "פרטנר" in file_name:
        return 'סה"כ צפיות'
    if "יס" in file_name:
        return "כמות הזמנות"
    if "screenil" in file_name_lower:
        return "Sessions"
    if ("סטינג" in file_name) and ("vod" in file_name_lower or "VOD" in file_name):
        return "כמות צופים"
    if "סטינג" in file_name:
        return "כמות צופים"
    if "סלקום" in file_name:
        return "כמות הזמנות"
    if ("הוט" in file_name) and ("ספריה" in file_name or "ספרייה" in file_name):
        return "סהכ הזמנות VOD"
    if "NEXT" in file_name_upper:
        return "כמות הזמנות"
    return None


def _decide_platform_label(file_name: str) -> Optional[str]:
    file_name_lower = file_name.lower()
    file_name_upper = file_name.upper()

    if "פרטנר" in file_name:
        return "פרטנר"
    if "יס" in file_name:
        return "YES"
    if "screenil" in file_name_lower:
        return "ScreenIL"
    if ("סטינג" in file_name) and ("vod" in file_name_lower or "VOD" in file_name):
        return "YES"
    if "סטינג" in file_name:
        return "YES"
    if "סלקום" in file_name:
        return "סלקום"
    if ("הוט" in file_name) and ("ספריה" in file_name or "ספרייה" in file_name):
        return "HOT"
    if "NEXT" in file_name_upper:
        return "NEXT"
    return None


def fill_template_files(
    mapped_excels: Dict[str, bytes],
    template_excel_bytes: bytes,
    prev_month: Optional[str] = None,
) -> Dict[str, bytes]:
    """
    For each mapped file, fill a fresh copy of the template (preserving formatting) and return:
      template_{mapped_filename}.xlsx bytes
    """
    outputs: Dict[str, bytes] = {}
    HEADER_ROW = 4
    prev_month_str_val = prev_month or previous_month_str()

    for mapped_name, mapped_bytes in mapped_excels.items():
        if not mapped_name.lower().endswith((".xlsx", ".xls")):
            continue

        mapped_df = pd.read_excel(BytesIO(mapped_bytes))
        mapped_df.columns = mapped_df.columns.astype(str).str.strip()

        program_col = _decide_template_program_col(mapped_name)
        if program_col is None:
            continue

        viewers_col = _decide_viewers_col(mapped_name)
        if viewers_col is None:
            continue

        platform = _decide_platform_label(mapped_name)
        mapped_house_col = _decide_mapped_house_col(mapped_name)

        # column existence checks like your script
        if program_col not in mapped_df.columns:
            continue
        if mapped_house_col not in mapped_df.columns:
            continue
        if viewers_col not in mapped_df.columns:
            continue

        # Load a fresh copy of template
        wb = load_workbook(BytesIO(template_excel_bytes))
        ws = wb.active

        # Set B1 and B2
        ws["B1"] = prev_month_str_val
        if platform is not None:
            ws["B2"] = platform

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Highlight platform in Q5:Q9
        for row in ws["Q5:Q9"]:
            for cell in row:
                if cell.value and str(cell.value).strip() == str(platform).strip():
                    cell.fill = yellow_fill

        # Highlight month in P5:P16
        for row in ws["P5:P16"]:
            for cell in row:
                if not cell.value:
                    continue
                val = str(cell.value).strip()
                if val == prev_month_str_val:
                    cell.fill = yellow_fill

        # Map headers row 4
        header_cells = ws[HEADER_ROW]
        header_map = {
            str(cell.value).strip(): cell.column
            for cell in header_cells
            if cell.value is not None
        }

        template_prog_col_name = "שם תוכנית בפלטפורמה"
        template_house_col_name = "מספר האוס בקשת TV"
        template_viewers_col_name = "כמות צפיות"
        template_date_col_name = "תאריך"

        required = [
            template_prog_col_name,
            template_house_col_name,
            template_viewers_col_name,
            template_date_col_name,
        ]
        if any(r not in header_map for r in required):
            continue

        prog_col_idx = header_map[template_prog_col_name]
        house_col_idx = header_map[template_house_col_name]
        viewers_col_idx = header_map[template_viewers_col_name]
        date_col_idx = header_map[template_date_col_name]

        start_row = HEADER_ROW + 1

        prog_values = mapped_df[program_col].tolist()
        house_values = mapped_df[mapped_house_col].tolist()
        viewers_values = mapped_df[viewers_col].tolist()

        row_idx = start_row
        for prog, house, viewers in zip(prog_values, house_values, viewers_values):
            if pd.isna(prog) and pd.isna(house) and pd.isna(viewers):
                continue

            ws.cell(row=row_idx, column=prog_col_idx, value=str(prog))
            ws.cell(row=row_idx, column=house_col_idx, value=house)
            ws.cell(row=row_idx, column=viewers_col_idx, value=viewers)
            ws.cell(row=row_idx, column=date_col_idx, value=prev_month_str_val)
            row_idx += 1

        out_name = f"template_{mapped_name}"
        out_buf = BytesIO()
        wb.save(out_buf)
        out_buf.seek(0)
        outputs[out_name] = out_buf.read()

    return outputs


# -----------------------------
# Orchestrator: run all + zip
# -----------------------------
@dataclass
class RunResult:
    zip_bytes: bytes
    summary: str


def run_pipeline_and_zip(
    platform_files: List[Tuple[str, bytes]],
    db_excel_bytes: bytes,
    template_excel_bytes: bytes,
    include_intermediate: bool = False,
    month_str: Optional[str] = None,   # NEW
) -> RunResult:
    cleaned = clean_input_files_to_excels(platform_files)
    mapped = map_house_numbers(cleaned, db_excel_bytes)
    templated = fill_template_files(mapped, template_excel_bytes, prev_month=month_str)  # CHANGED

    # Build ZIP
    zip_buf = BytesIO()
    with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as z:
        for name, b in templated.items():
            z.writestr(name, b)

        if include_intermediate:
            for name, b in cleaned.items():
                z.writestr(name, b)
            for name, b in mapped.items():
                z.writestr(name, b)

    zip_buf.seek(0)

    summary = (
        f"Cleaned files: {len(cleaned)}\n"
        f"Mapped files: {len(mapped)}\n"
        f"Template outputs: {len(templated)}\n"
    )

    return RunResult(zip_bytes=zip_buf.read(), summary=summary)
